import os
import time
import pandas as pd
from dataclasses import dataclass
from typing import List, Dict
from flask import jsonify, request, render_template, send_from_directory
from openpyxl.styles import Alignment
from werkzeug.utils import secure_filename
from flask_socketio import SocketIO
from openpyxl.utils import get_column_letter

from pyFCM import fuzzy_match_EM, extract_specifications, init_atlas, fuzzy_match

# 配置上传文件存储路径
UPLOAD_FOLDER = 'uploads/estimation_water'
OUTPUT_FOLDER = 'output/estimation_water'
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
UNIT_MAPPING_LEN_MM = {"米": 1000, "m": 1000, "dm": 100, "cm": 10, "mm": 1}

# 确保上传目录和输出目录存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)


# socketio = SocketIO() 这里似乎没必要单独初始化这个定义，def init_routes(app, socketio): 这里作为参数传入了一个socketio


@dataclass
class EquipmentMaterial:
    """设备材料信息类"""
    name: str  # 名称
    specification: str  # 规格
    material: str  # 材料
    unit: str  # 单位
    quantity: float  # 数量
    remarks: str  # 备注


Atlas_PipeFittingsQ235A = {}  # 钢管配件重量表
Atlas_PipeFittingsDuctileIron = {}  # 球墨铸铁管配件
Atlas_Valve = {}  # 阀门价格表
Atlas_Equipment = {}


# 读取重量表
def atlas():
    # 读取 Excel 文件中的所有工作表
    tXls = pd.ExcelFile("templates/250407管配件重量表.xlsx")
    for feSheetName in tXls.sheet_names:
        df = pd.read_excel(tXls, sheet_name=feSheetName, header=0, index_col=0)
        # 创建一个字典，用于存储每种管配件的每米重量
        # 遍历表格，提取每米重量
        for _, feRow in df.iterrows():
            # 遍历 feRow 中的每一个单元格
            dn1, dn2 = 0, 0
            for feColName, feValue in feRow.items():
                if feColName == "管径1":
                    dn1 = int(feValue)
                elif feColName == "管径2":
                    dn2 = int(feValue)
                else:
                    if pd.notna(feValue):  # 使用 pd.notna() 判断值是否不是 NaN
                        if feSheetName == "Q235A":
                            # 如果 column_name 不在 Atlas_PipeFittingsQ235A 中，初始化一个空字典
                            if feColName not in Atlas_PipeFittingsQ235A:
                                Atlas_PipeFittingsQ235A[feColName] = {}
                            # 如果 dn1 不在 Atlas_PipeFittingsQ235A[column_name] 中，初始化一个空字典
                            if dn1 not in Atlas_PipeFittingsQ235A[feColName]:
                                Atlas_PipeFittingsQ235A[feColName][dn1] = {}
                            # 存储 dn2 和对应的重量
                            Atlas_PipeFittingsQ235A[feColName][dn1][dn2] = feValue
                        elif feSheetName == "球铁":
                            # 如果 column_name 不在 Atlas_PipeFittingsDuctileIron 中，初始化一个空字典
                            if feColName not in Atlas_PipeFittingsDuctileIron:
                                Atlas_PipeFittingsDuctileIron[feColName] = {}
                            # 如果 dn1 不在 Atlas_PipeFittingsDuctileIron[column_name] 中，初始化一个空字典
                            if dn1 not in Atlas_PipeFittingsDuctileIron[feColName]:
                                Atlas_PipeFittingsDuctileIron[feColName][dn1] = {}
                            # 存储 dn2 和对应的重量
                            Atlas_PipeFittingsDuctileIron[feColName][dn1][dn2] = feValue

    # 读取阀门价格表
    df = pd.read_excel("templates/250410阀门.xlsx", header=0, index_col=0)
    for _, feRow in df.iterrows():
        dn1 = 0
        for feColName, feValue in feRow.items():
            if feColName == "管径1":
                dn1 = int(feValue)
            else:
                if pd.notna(feValue):
                    Atlas_Valve[feColName] = {dn1: feValue}
    # 读取设备价格表
    df = pd.read_excel("templates/250410设备.xlsx", header=0, index_col=0)
    for _, feRow in df.iterrows():
        tName = ""
        for feColName, feValue in feRow.items():
            if pd.notna(feValue):
                if feColName == "设备名称":
                    tName = feValue
                    Atlas_Equipment[tName] = {}
                else:
                    Atlas_Equipment[tName][feColName] = feValue
    # print(Atlas_Equipment.keys())
    init_atlas(Atlas_PipeFittingsQ235A, Atlas_PipeFittingsDuctileIron, Atlas_Valve, Atlas_Equipment)


def allowed_file(pFilename):
    """检查文件类型是否允许"""
    return '.' in pFilename and pFilename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def safe_filename(pFilename: str) -> str:
    tName, tExt = os.path.splitext(pFilename)  # 获取文件扩展名
    rtSafeName = tName.replace('/', '_').replace('\\', '_')  # 替换不安全的字符，但保留中文
    return rtSafeName + tExt  # 重新组合文件名和扩展名


def _stage_update(pPercent, pStage, pSessionId=None, pSocketio=None):
    if pSocketio:
        pSocketio.emit('progress', {'progress': pPercent, 'stage': pStage, 'sessionId': pSessionId})
    time.sleep(0)


def process_excel_file(pFilePath: str, pSessionId: str, pSocketio=None) -> Dict[str, List[EquipmentMaterial]]:
    """
    处理Excel文件，提取设备材料表信息
    Args:
        pFilePath: Excel文件路径
        pSessionId: 会话ID，用于更新进度
        pSocketio: SocketIO实例，用于进度更新
    Returns:
        以所属单体为key，设备材料列表为value的字典
    """
    _TARGET_WORDS = {"序号": ["序号", "编号"],
                     "所属单体": ["单体", "所属单体", "构筑物", "位置", "设备位置", "安装位置", "安装地点"],
                     "名称": ["名称", "设备名称"],
                     "规格": ["规格", "规格尺寸", "规格参数", "型号规格"],
                     "材料": ["材料", "材质"],
                     "单位": ["单位"],
                     "数量": ["数量"],
                     "备注": ["备注"]}
    # 假设表头包含这些字段
    # _Required_Columns = ["序号", "所属单体", "名称", "规格", "材料", "单位", "数量", "备注"]
    _REQUIRED_COLUMNS = ["名称", "规格", "单位", "数量"]

    def _match_row(pRow):
        rtTargetCol = {}
        tTargetSim = {}
        for feKey in _TARGET_WORDS:
            tTargetSim[feKey] = 0.0
        for feCol in range(0, len(pRow)):

            feCellValue = pRow[feCol]
            if pd.isna(feCellValue):
                continue
            # print(feCol, feCellValue)
            tMatchKey = None
            tMatchSim = 0.0
            for feKey in _TARGET_WORDS.keys():  # 与目标字段逐个匹配
                tMatchedWord, tSimilarityScore = fuzzy_match(feCellValue, _TARGET_WORDS[feKey])
                # print("match", _matched_word, _similarity_score)
                if tSimilarityScore[0] > tMatchSim:
                    tMatchSim = tSimilarityScore[0]
                    tMatchKey = feKey
            # print("ks", feCellValue, _match_key, _match_sim)
            if tMatchKey is not None:
                if tMatchSim > max(tTargetSim[tMatchKey], 0.8):
                    rtTargetCol[tMatchKey] = feCol
                    tTargetSim[tMatchKey] = tMatchSim

        return rtTargetCol

    try:

        # 使用 openpyxl 引擎读取 Excel 文件
        tExcelFile = pd.ExcelFile(pFilePath, engine='openpyxl')

        rtDict = {}
        tSheetsCount = len(tExcelFile.sheet_names)

        # 遍历所有工作表
        for feSheetIndex, feSheetName in enumerate(tExcelFile.sheet_names):
            progress = 10 + (feSheetIndex / tSheetsCount) * 60  # 10-70% 的进度
            _stage_update(progress, f'正在处理工作表: {feSheetName}……', pSessionId, pSocketio)

            tDFSheet = pd.read_excel(tExcelFile, sheet_name=feSheetName, engine='openpyxl', header=None)
            tMatchHeadRow = 0
            tCurrentRow = 0
            _key_exchange = {}
            for feRowIndex, feRow in tDFSheet.iterrows():
                tTargetCol = _match_row(feRow)
                tMatchHeadRow = feRowIndex
                _key_exchange = tTargetCol
                tCurrentRow = tCurrentRow + 1
                if all(col in _key_exchange.keys() for col in _REQUIRED_COLUMNS) or tCurrentRow > 10:
                    break

            print("tMatchHeadRow", tMatchHeadRow)
            print(_key_exchange)

            tDFSheet = pd.read_excel(tExcelFile, sheet_name=feSheetName, engine='openpyxl', header=tMatchHeadRow)
            if all(col in _key_exchange.keys() for col in _REQUIRED_COLUMNS):  # 判断要求的字段是否都在识别结果中能找到
                # 找到目标表格
                tLastIndividual = ""
                tLastEM = None

                for feRowIndex, (_, feRow) in enumerate(tDFSheet.iterrows()):
                    if "所属单体" not in _key_exchange.keys():
                        tIndividual = (feSheetName.replace("gpj", "").replace("GPJ", "").replace("sb", "")
                                       .replace("SB", "").replace("管配件", "").replace("设备", "").replace(" ", ""))
                    else:
                        # 处理单体单元格合并情况
                        tIndividual = feRow.iloc[_key_exchange["所属单体"]]
                    if pd.isna(tIndividual):
                        if tLastIndividual == "":
                            continue
                        else:
                            tIndividual = tLastIndividual
                    else:
                        tLastIndividual = tIndividual
                    # 处理名称单元格合并情况

                    tEMname = feRow.iloc[_key_exchange["名称"]]
                    if pd.isna(tEMname):
                        if pd.isna(feRow.iloc[_key_exchange["数量"]]):
                            if tLastEM is not None:
                                tLastEM.specification += str(feRow.iloc[_key_exchange["规格"]])
                            continue
                        else:
                            if tLastEM is not None:
                                tEMname = tLastEM.name
                            else:
                                continue

                    tQList = str(feRow.iloc[_key_exchange["数量"]]).split('/')  # 如果有多个规格写在同一行的，例：墙管 DN500/DN300 个 40/91
                    tSpStr = str(feRow.iloc[_key_exchange["规格"]])
                    tSpList = []
                    if len(tQList) > 1:
                        tSpList = tSpStr.split('/')
                    else:
                        tSpList.append(str(feRow.iloc[_key_exchange["规格"]]))

                    for i in range(0, len(tQList)):
                        tQ = tQList[i]
                        if i <= len(tSpList) - 1:
                            tSp = tSpList[i]
                        else:
                            tSp = tSpList[len(tSpList) - 1]
                        tMaterial = ""
                        if "材料" in _key_exchange.keys():
                            if pd.notna(feRow.iloc[_key_exchange["材料"]]):
                                tMaterial = str(feRow.iloc[_key_exchange["材料"]])
                        tRemark = ""
                        if "备注" in _key_exchange.keys():
                            if pd.notna(feRow.iloc[_key_exchange["备注"]]):
                                tRemark = str(feRow.iloc[_key_exchange["备注"]])
                        if pd.notna(feRow.iloc[_key_exchange["单位"]]):
                            tUnit = str(feRow.iloc[_key_exchange["单位"]])
                        else:
                            tUnit = ""
                        # 创建设备材料对象
                        tEM = EquipmentMaterial(
                            name=str(tEMname),
                            specification=str(tSp),
                            material=tMaterial,
                            unit=tUnit,
                            quantity=tQ if (pd.notna(tQ) and tQ != "nan") else 0.0,
                            remarks=tRemark
                        )
                        # 将设备材料添加到对应单体的列表中
                        if tIndividual not in rtDict:
                            rtDict[tIndividual] = []
                        rtDict[tIndividual].append(tEM)
                        tLastEM = tEM

        return rtDict

        # raise ValueError("未找到符合要求的设备材料表")

    except Exception as e:
        _stage_update(0, f"处理Excel文件时出错: {str(e)}", pSessionId, pSocketio)
        raise Exception(f"处理Excel文件时出错: {str(e)}")


def write_to_excel(pEMDict: Dict[str, List[EquipmentMaterial]], pOriginalFilename: str) -> str:
    # 复制单元格格式
    def _copy_cell_format(pCell, pTemplateCell):
        if pTemplateCell.has_style:
            pCell.font = pTemplateCell.font.copy()  # 复制字体
            pCell.border = pTemplateCell.border.copy()  # 复制边框
            pCell.fill = pTemplateCell.fill.copy()  # 复制填充
            pCell.number_format = pTemplateCell.number_format  # 复制数字格式
            pCell.protection = pTemplateCell.protection.copy()  # 复制保护
            pCell.alignment = pTemplateCell.alignment.copy()  # 复制对齐方式


    _individual_sum_row = {}
    # 写入总表
    def write_to_excle_summary():
        def cell_format(pWorkSheet, pTemplateWS, pCurrentRow, pNameColIdx, pSumColIdx, pValue):
            tCell = pWorkSheet.cell(row=pCurrentRow, column=pNameColIdx)
            tCell.value = pValue
            tCell.alignment = Alignment(horizontal='right', vertical='center')
            tCell = pWorkSheet.cell(row=pCurrentRow, column=pSumColIdx)
            tCell.value = f"=SUM(D{pCurrentRow}:G{pCurrentRow})"
            row_formate(pWorkSheet, pTemplateWS, pCurrentRow)

        def row_formate(pWorkSheet, pTemplateWS, pRow):
            pWorkSheet.row_dimensions[pRow].height = pTemplateWS.row_dimensions[7].height  # 调整行高
            for feCol in range(1, 13):
                tCell = pWorkSheet.cell(row=pRow, column=feCol)
                tCell.border = pTemplateWS.cell(row=8, column=feCol).border.copy()  # 使用模板的边框样式
                if feCol != 3:
                    tCell.alignment = pTemplateWS.cell(row=8, column=feCol).alignment.copy()  # 使用模板的对齐方式

        tWorkSheet = writer.sheets['总表']  # 获取工作表对象
        # 复制模板的前7行格式
        for feRow in range(1, 8):  # 复制前7行
            for feCol in range(1, len(template_ws[1]) + 1):
                tTemplateCell = template_ws.cell(row=feRow, column=feCol)  # 获取模板单元格
                tCell = tWorkSheet.cell(row=feRow, column=feCol)  # 获取目标单元格
                tCell.value = tTemplateCell.value  # 复制单元格值
                _copy_cell_format(tCell, tTemplateCell)  # 复制单元格格式

        # 复制合并单元格
        for feMergedRange in template_ws.merged_cells.ranges:
            if feMergedRange.min_row <= 7:  # 只复制前7行的合并单元格
                tWorkSheet.merge_cells(str(feMergedRange))
        # 复制列宽
        for feCol in range(1, len(template_ws[1]) + 1):
            tColLetter = get_column_letter(feCol)
            tWorkSheet.column_dimensions[tColLetter].width = template_ws.column_dimensions[tColLetter].width
            # print(column_letter, template_ws.column_dimensions[column_letter].width)
        # 复制行高
        for feRow in range(1, 8):  # 复制前7行的行高
            tWorkSheet.row_dimensions[feRow].height = template_ws.row_dimensions[feRow].height
        # 找到"工程或费用名称"列的索引
        tNameColIdx = 3
        tSumColIdx = 8
        # 从第8行开始写入数据
        tCurrentRow = 8

        # 写入每个key，并在key之间添加3个空行
        for feIndivName in pEMDict.keys():
            # 写入key
            tCell = tWorkSheet.cell(row=tCurrentRow, column=tNameColIdx)
            tCell.value = feIndivName
            tCell = tWorkSheet.cell(row=tCurrentRow, column=tSumColIdx)
            tCell.value = f"=SUM(D{tCurrentRow}:G{tCurrentRow})"
            row_formate(tWorkSheet, template_ws, tCurrentRow)
            tCurrentRow += 1
            cell_format(tWorkSheet, template_ws, tCurrentRow, tNameColIdx, tSumColIdx, "土建")
            tCurrentRow += 1
            cell_format(tWorkSheet, template_ws, tCurrentRow, tNameColIdx, tSumColIdx, "管配件")
            tCell = tWorkSheet.cell(row=tCurrentRow, column=5)
            tCell.value = f"=ROUND('{feIndivName}gpj'!H{_individual_sum_row[feIndivName + 'gpj']}/10000,2)"
            tCurrentRow += 1
            cell_format(tWorkSheet, template_ws, tCurrentRow, tNameColIdx, tSumColIdx, "设备")
            tCell = tWorkSheet.cell(row=tCurrentRow, column=5)
            tCell.value = f"=ROUND('{feIndivName}sb'!H{_individual_sum_row[feIndivName + 'sb']}/10000,2)"
            for feCol in range(4, 8):
                tCell = tWorkSheet.cell(row=tCurrentRow - 3, column=feCol)
                tCell.value = f"=SUM({get_column_letter(feCol)}{tCurrentRow - 2}:{get_column_letter(feCol)}{tCurrentRow})"

            tCurrentRow += 1

        for feCol in range(4, 8):
            tCell = tWorkSheet.cell(row=7, column=feCol)
            tCell.value = f"=SUM({get_column_letter(feCol)}{7 + 1}:{get_column_letter(feCol)}{tCurrentRow})"
        tCell = tWorkSheet.cell(row=7, column=8)
        tCell.value = "=SUM(D7:G7)"

    # 写入单项概算
    def write_to_excel_individual():
        def find_closest_key(pRandNum, pDict):
            """
            找到与随机整数差值最小的字典键。
            参数:
                pRandNum (int): 随机整数
                dictionary (dict): 字典，键是整数
            返回:
                int: 与随机整数差值最小的键
            """
            # 初始化最小差值和最接近的键
            min_diff = float('inf')  # 设置为无穷大
            closest_key = None

            # 遍历字典的键
            for feKey in pDict.keys():
                # 计算差值
                diff = abs(feKey - pRandNum)
                # 更新最小差值和最接近的键
                if diff < min_diff:
                    min_diff = diff
                    closest_key = feKey

            return closest_key

        # =============================================设备材料表================================================================
        category = {"gpj": ["管配件", "材料"], "sb": ["设备"]}
        for feIndivName in pEMDict.keys():
            for feSuffix in category.keys():
                # 创建第二个工作表（设备材料表）
                template_ws2 = template_wb['Sheet2']  # 获取模板的第二个工作表
                worksheet2 = workbook.create_sheet(feIndivName + feSuffix)  # 创建新的工作表
                # 复制Sheet2的前7行格式
                for feRow in range(1, 8):
                    for feCol in range(1, len(template_ws2[1]) + 1):
                        tTemplateCell = template_ws2.cell(row=feRow, column=feCol)  # 获取模板单元格
                        tCell = worksheet2.cell(row=feRow, column=feCol)  # 获取目标单元格

                        # 复制单元格值
                        if feRow == 3 and feCol == 2:
                            tCell.value = feIndivName + " " + category[feSuffix][0]
                        else:
                            tCell.value = tTemplateCell.value
                        _copy_cell_format(tCell, tTemplateCell)  # 复制单元格格式

                # 复制Sheet2的合并单元格
                for feMergedRange in template_ws2.merged_cells.ranges:
                    if feMergedRange.min_row <= 7:  # 只复制前7行的合并单元格
                        worksheet2.merge_cells(str(feMergedRange))

                # 复制Sheet2的列宽
                for feCol in range(1, len(template_ws2[1]) + 1):
                    tColLetter = get_column_letter(feCol)
                    if tColLetter in template_ws2.column_dimensions:
                        worksheet2.column_dimensions[tColLetter].width = template_ws2.column_dimensions[tColLetter].width

                # 复制Sheet2的行高
                for feRow in range(1, 8):  # 复制前7行的行高
                    worksheet2.row_dimensions[feRow].height = template_ws2.row_dimensions[feRow].height

                tCurrentRow = 8

                for feEM in pEMDict[feIndivName]:
                    tBM, tFlange, tMaterial, tScore, tType, tResult = fuzzy_match_EM(feEM)
                    dn1, dn2 = 0, 0
                    tValue = ""
                    tPrice, tPriceFlange = 1, 1
                    tDensity = ""  # 与铁的容重比
                    tAtlas = Atlas_PipeFittingsQ235A
                    if tMaterial in ["Q235A", "Q235B", "Q235C", "Q235D", "Q235E", "钢"]:
                        tPrice, tPriceFlange = 1, 1
                        tDensity = ""
                        tAtlas = Atlas_PipeFittingsQ235A
                    elif tMaterial in ["SS304"]:
                        tPrice, tPriceFlange = 3, 1
                        tDensity = "*7.93/7.85"
                        tAtlas = Atlas_PipeFittingsQ235A
                    elif tMaterial in ["SS316"]:
                        tPrice, tPriceFlange = 5, 1
                        tDensity = "*8.0/7.85"
                        tAtlas = Atlas_PipeFittingsQ235A
                    elif tMaterial in ["球铁"]:
                        tPrice, tPriceFlange = 7, 0
                        tAtlas = Atlas_PipeFittingsDuctileIron

                    if len(tResult["管径"]) > 0:
                        dn1 = tResult["管径"][0]
                        dn2 = dn1
                        if len(tResult["管径"]) > 1:
                            dn2 = tResult["管径"][1]
                    if tType == "管配件" and tScore > 0:
                        if tBM in tAtlas.keys():
                            tDic = tAtlas[tBM][find_closest_key(dn1, tAtlas[tBM])]
                            tFlangeDn1 = find_closest_key(dn1, tAtlas["法兰"])
                            tFlangeWeight = tAtlas["法兰"][tFlangeDn1][tFlangeDn1]
                            tCircleStr, tLengthStr = "", ""

                            if tBM in ["直管", "套管",
                                       "穿墙套管", "柔性防水套管A型Ⅰ型", "柔性防水套管A型Ⅱ型", "柔性防水套管B型Ⅰ型",
                                       "柔性防水套管B型Ⅱ型", "法兰套管A型Ⅰ型", "法兰套管A型Ⅱ型", "法兰套管B型Ⅰ型",
                                       "法兰套管B型Ⅱ型", "刚性防水套管A型", "刚性防水套管B型", "刚性防水套管C型"]:
                                if feEM.unit not in UNIT_MAPPING_LEN_MM.keys():
                                    tStr = tResult["长度"]
                                    if float(tStr) == 0.0:
                                        tStr = "300"
                                    tLengthStr = "*" + str(tStr) + "/1000"
                                if tBM == "穿墙套管":
                                    tCircle = Atlas_PipeFittingsQ235A["穿墙套管-配件"][
                                        find_closest_key(dn1, Atlas_PipeFittingsQ235A["穿墙套管-配件"])]
                                    tCircleStr = "+" + str(tCircle[find_closest_key(dn2, tCircle)])

                            tValue = (
                                f"=({tDic[find_closest_key(dn2, tDic)]}{tLengthStr}{tCircleStr})/1000{tDensity}*K{tPrice}"
                                f"+{tFlange}*{tFlangeWeight}/1000{tDensity}*K{tPrice + tPriceFlange}")

                        if tType == "阀门" and tResult["功率"] == 0.0:
                            if len(tResult["管径"]) > 0:
                                if tResult["管径"][0] >= 600:
                                    tType = "设备"
                                else:
                                    tType = "材料"
                            else:
                                tType = "材料"
                    if tType == "":
                        tType = "材料"
                    if tType not in category[feSuffix]:
                        continue
                    tCell = worksheet2.cell(row=tCurrentRow, column=2)
                    tCell.value = tCurrentRow - 7
                    tCell = worksheet2.cell(row=tCurrentRow, column=3)
                    tCell.value = feEM.name
                    tCell = worksheet2.cell(row=tCurrentRow, column=4)
                    if feEM.material != "nan":
                        tCell.value = f"{feEM.specification} {feEM.material}"
                    else:
                        tCell.value = feEM.specification
                    tCell = worksheet2.cell(row=tCurrentRow, column=5)
                    tCell.value = feEM.unit
                    tCell = worksheet2.cell(row=tCurrentRow, column=6)
                    tCell.value = feEM.quantity
                    tCell = worksheet2.cell(row=tCurrentRow, column=7)
                    tCell.value = str(tValue)
                    tCell = worksheet2.cell(row=tCurrentRow, column=8)
                    tCell.value = f"=F{tCurrentRow}*G{tCurrentRow}"

                    tCell = worksheet2.cell(row=tCurrentRow, column=13)
                    tCell.value = tBM
                    tCell = worksheet2.cell(row=tCurrentRow, column=12)
                    tCell.value = tScore
                    tCell = worksheet2.cell(row=tCurrentRow, column=14)
                    if tResult["长度"] != 0:
                        tCell.value = f"DN{dn1}×DN{dn2} L=" + str(tResult["长度"]) + str(tResult["长度单位"])
                    else:
                        tCell.value = f"DN{dn1}×DN{dn2}"

                    tCell = worksheet2.cell(row=tCurrentRow, column=17)
                    tCell.value = str(tType)

                    worksheet2.row_dimensions[tCurrentRow].height = template_ws2.row_dimensions[8].height
                    for feCol in range(1, 9):
                        tCell = worksheet2.cell(row=tCurrentRow, column=feCol)
                        tTemplateCell = template_ws2.cell(row=8, column=feCol)
                        _copy_cell_format(tCell, tTemplateCell)  # 复制单元格格式

                    tCurrentRow += 1

                for feRow in range(tCurrentRow, tCurrentRow + 7):
                    worksheet2.row_dimensions[feRow].height = template_ws2.row_dimensions[23].height
                    for feCol in range(1, 9):
                        tCell = worksheet2.cell(row=feRow, column=feCol)
                        tTemplateCell = template_ws2.cell(row=feRow - tCurrentRow + 23, column=feCol)
                        _copy_cell_format(tCell, tTemplateCell)  # 复制单元格格式
                        if feCol in [3, 4]:
                            tCell.value = tTemplateCell.value
                        elif feCol == 5:
                            if feRow - tCurrentRow not in [0, 6]:
                                tCell.value = "元"
                        elif feCol == 8:
                            if feRow - tCurrentRow == 1:
                                tCell.value = f"=SUM(H8:H{feRow - 1})"
                            elif feRow - tCurrentRow in [2, 4]:
                                tCell.value = f"=H{feRow - 1}*D{feRow}"
                            elif feRow - tCurrentRow in [3, 5]:
                                tCell.value = f"=SUM(H{feRow - 2}:H{feRow - 1})"
                tCell = worksheet2.cell(row=4, column=2)
                tCell.value = f'="估算价值(元)："&ROUND(H{tCurrentRow + 5},0)'
                _individual_sum_row[feIndivName + feSuffix] = tCurrentRow + 5
    #  函数本体从这里开始执行
    try:
        tTimeStamp = time.strftime("%y%m%d%H%M%S")  # 生成时间戳格式的文件名 (YYMMDDHHMMSS)
        tSafeFilename = safe_filename(pOriginalFilename)  # 使用安全的文件名处理函数，但保留原始字符
        tNewFilename = f"{tTimeStamp}_{tSafeFilename}"  # 构建新文件名
        tOutputPath = os.path.join(OUTPUT_FOLDER, tNewFilename)

        tTemplatePath = os.path.join('templates', 'standard.xlsx')  # 读取标准模板
        # 使用openpyxl直接读取模板文件以保留格式
        from openpyxl import load_workbook
        template_wb = load_workbook(tTemplatePath)
        template_ws = template_wb['Sheet1']

        df = pd.DataFrame()  # 创建一个空的DataFrame作为初始工作表
        with pd.ExcelWriter(tOutputPath, engine='openpyxl') as writer:  # 创建一个新的Excel写入器
            df.to_excel(writer, sheet_name='总表', index=False)  # 先写入空的DataFrame以创建工作表
            workbook = writer.book  # 获取工作簿
            write_to_excel_individual()
            write_to_excle_summary()
        return tOutputPath

    except Exception as e:
        raise Exception(f"写入Excel文件时出错: {str(e)}")


def init_routes(app, socketio):
    """初始化路由和Socket.IO事件处理"""

    @app.route('/estimation/water')
    def estimation_water():
        """水厂智能估算页面路由"""
        atlas()
        return render_template('estimation_water.html')

    @app.route('/estimation/water/upload', methods=['POST'])
    def upload_water_file():
        """处理文件上传"""
        if 'file0' not in request.files:
            return jsonify({'error': '没有文件被上传'}), 400

        file = request.files['file0']
        if file.filename == '':
            return jsonify({'error': '没有选择文件'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': '不支持的文件类型'}), 400

        # 获取会话ID
        session_id = request.headers.get('X-Session-ID')

        try:
            # 使用安全的文件名处理函数，但保留原始字符
            filename = safe_filename(file.filename)
            _stage_update(5, f'文件 {file.filename} 上传成功！', session_id, socketio)

            # 添加时间戳确保文件名唯一
            timestamp = time.strftime("%Y%m%d_%H%M%S")
            new_filename = f"{timestamp}_{filename}"
            file_path = os.path.join(UPLOAD_FOLDER, new_filename)

            file.save(file_path)  # 保存文件
            _stage_update(10, f'文件 {file.filename} 保存成功！，数据读取中……', session_id, socketio)  # 发送进度更新

            equipment_dict = process_excel_file(file_path, session_id, socketio)  # 处理Excel文件，传入session_id
            # print("读取完成！")
            _stage_update(70, f'文件 {file.filename} 读取完成，正在生成结果文件……', session_id, socketio)  # 发送进度更新

            output_path = write_to_excel(equipment_dict, file.filename)  # 使用原始文件名，写入新的Excel文件

            _stage_update(100, f'处理完成，已输出文件: {os.path.basename(output_path)}', session_id, socketio)  # 发送进度更新

            return jsonify({
                'message': '文件上传成功',
                'filename': new_filename,
                'progress': 100,
                'output_file': os.path.basename(output_path),
                'equipment_data': {
                    unit: [
                        {
                            'name': eq.name,
                            'specification': eq.specification,
                            'material': eq.material,
                            'unit': eq.unit,
                            'quantity': eq.quantity,
                            'remarks': eq.remarks
                        }
                        for eq in equipment_list
                    ]
                    for unit, equipment_list in equipment_dict.items()
                }
            })

        except Exception as e:
            _stage_update(0, f'上传失败: {str(e)}', session_id, socketio)  # 发送错误信息
            return jsonify({'error': str(e)}), 500

    @app.route('/estimation/water/download/<filename>')
    def download_file(filename):
        """处理文件下载"""
        try:
            return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    @socketio.on('upload')
    def handle_upload(data):
        """处理上传开始事件"""
        filename = data.get('filename')
        session_id = data.get('sessionId')
        print(f'Upload started for file: {filename} (Session: {session_id})')
