import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy

def locate_sheet(workbook, sheet_name):
    """定位到指定工作簿中的工作表"""
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    else:
        raise ValueError(f"Worksheet '{sheet_name}' not found in the workbook")

def sheet_to_numeric(sheet):
    """将工作表中的所有数据转换为数值格式"""
    for row in sheet.iter_rows():
        for cell in row:
            try:
                if isinstance(cell.value, str):
                    cell.value = float(cell.value)
            except ValueError:
                continue

def find_first_and_last_row(sheet, start_col):
    """找到建筑工程或其他相应列的首个非空单元格和最后一个带边框单元格的行号"""
    start_row = None
    last_row = None
    for row in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=start_col)
        if cell.value is not None and start_row is None:
            start_row = row
        if cell.border and (cell.border.left.style or cell.border.right.style or cell.border.top.style or cell.border.bottom.style):
            last_row = row
    return start_row, last_row

def copy_column(sheet_from, sheet_to, from_col, to_col, start_row, end_row):
    """复制整列数据从源工作表到目标工作表，保留原格式和单元格边框"""
    for row in range(start_row, end_row + 1):
        from_cell = sheet_from.cell(row=row, column=from_col)
        if isinstance(from_cell, openpyxl.cell.cell.MergedCell):
            continue  # 跳过合并的单元格
        to_cell = sheet_to.cell(row=row, column=to_col)
        to_cell.value = from_cell.value
        to_cell.font = copy(from_cell.font)
        to_cell.fill = copy(from_cell.fill)
        to_cell.number_format = from_cell.number_format
        to_cell.protection = copy(from_cell.protection)
        to_cell.alignment = copy(from_cell.alignment)
        to_cell.border = copy(from_cell.border)

def process_field(sheet_from, sheet_to, t_field_data, p_field_data):
    """处理单个字段的复制操作"""
    if len(t_field_data) > 1:
        # 如果有多个坐标，将这些列相加后再复制
        start_row, end_row = find_first_and_last_row(sheet_from, t_field_data[0]['col'] + 1)
        start_row += 1  # 从坐标的 col+1 行开始复制
        summed_values = [0] * (end_row - start_row + 1)
        for t_data in t_field_data:
            from_col = t_data['col'] + 1
            for i, row in enumerate(range(start_row, end_row + 1)):
                cell_value = sheet_from.cell(row=row, column=from_col).value
                if isinstance(cell_value, (int, float)):
                    summed_values[i] += cell_value
        to_col = p_field_data[0]['col'] + 1
        for i, row in enumerate(range(start_row, end_row + 1)):
            to_cell = sheet_to.cell(row=row, column=to_col)
            to_cell.value = summed_values[i]
            # 保留原格式和边框
            from_cell = sheet_from.cell(row=row, column=t_field_data[0]['col'] + 1)
            to_cell.font = copy(from_cell.font)
            to_cell.fill = copy(from_cell.fill)
            to_cell.number_format = from_cell.number_format
            to_cell.protection = copy(from_cell.protection)
            to_cell.alignment = copy(from_cell.alignment)
            to_cell.border = copy(from_cell.border)
    else:
        # 如果只有一个坐标，直接复制
        t_data = t_field_data[0]
        p_data = p_field_data[0]
        start_row, end_row = find_first_and_last_row(sheet_from, t_data['col'] + 1)
        start_row += 1  # 从坐标的 col+1 行开始复制
        copy_column(sheet_from, sheet_to, t_data['col'] + 1, p_data['col'] + 1, start_row, end_row)


def table_format(test_file, standard_file, t_dict):
    p_dict = {'表单名称': '总表', '建筑工程': [{'row': 3, 'col': 2, 'sim': '1.0'}],
              '安装工程': [{'row': 3, 'col': 3, 'sim': '1.0'}],
              '设备及工器具购置费': [{'row': 3, 'col': 4, 'sim': '1.0'}],
              '其他费用': [{'row': 3, 'col': 5, 'sim': '1.0'}], '合计': [{'row': 3, 'col': 6, 'sim': '1.0'}],
              '单位': [{'row': 3, 'col': 7, 'sim': '1.0'}], '数量': [{'row': 3, 'col': 8, 'sim': '1.0'}],
              '单位价值元': [{'row': 3, 'col': 9, 'sim': '1.0'}], '序号': [{'row': 2, 'col': 0, 'sim': '1.0'}],
              '工程或费用名称': [{'row': 2, 'col': 1, 'sim': '1.0'}], '备注': [{'row': 10, 'col': 3, 'sim': '1.0'}]}
    # 加载工作簿
    print(test_file,standard_file)
    test_wb = openpyxl.load_workbook(test_file)
    standard_wb = openpyxl.load_workbook(standard_file)

    # 定位到表单名称对应的工作表
    test_sheet = locate_sheet(test_wb, t_dict['表单名称'])
    standard_sheet = locate_sheet(standard_wb, p_dict['表单名称'])

    # 将工作表数值化
    sheet_to_numeric(test_sheet)

    # 动态处理字典中的所有字段
    for field in t_dict:
        if field != '表单名称' and field in p_dict:
            process_field(test_sheet, standard_sheet, t_dict[field], p_dict[field])

    # 保存到新的标准文件
    standard_wb.save("standard_result.xlsx")


