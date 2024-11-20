import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment

import pandas as pd


def new_workbook():
    rt_workbook = dict()
    return rt_workbook


def get_workbook(p_file_path):
    _data_frame_dict = pd.read_excel(p_file_path, sheet_name=None, header=None)
    return _data_frame_dict


def trans_to_xlsx(p_file_path):
    border_styles = {
        0: "none",
        1: "thin",
        2: "medium",
        3: "dashed",
        4: "dotted",
        5: "thick",
        6: "double",
        7: "hair",
        8: "medium dashed",
        9: "thin dash-dotted",
        10: "medium dash-dotted",
        11: "thin dash-dot-dotted",
        12: "medium dash-dot-dotted",
        13: "slanted medium dash-dotted"
    }
    # 打开.xls文件
    workbook_xls = xlrd.open_workbook(p_file_path, formatting_info=True)

    # 创建一个新的.xlsx工作簿
    workbook_xlsx = Workbook()
    # 移除自动创建的默认工作表
    workbook_xlsx.remove(workbook_xlsx.active)

    # 读取.xls文件中的所有工作表
    for fe_index in range(workbook_xls.nsheets):
        worksheet_xls = workbook_xls.sheet_by_index(fe_index)
        worksheet_xlsx = workbook_xlsx.create_sheet(title=worksheet_xls.name)  # 创建一个新的工作表，并命名为.xls工作表的名称

        # 复制行
        for row_idx in range(worksheet_xls.nrows):
            for col_idx in range(worksheet_xls.ncols):
                cell_xls = worksheet_xls.cell(row_idx, col_idx)
                cell_xlsx = worksheet_xlsx.cell(row=row_idx + 1, column=col_idx + 1, value=cell_xls.value)


                # 检查是否有XF索引
                if cell_xls.xf_index is not None:
                    xf = workbook_xls.xf_list[cell_xls.xf_index]
                    # 复制边框样式
                    border = Border(left=Side(style=border_styles[xf.border.left_line_style]),
                                    right=Side(style=border_styles[xf.border.right_line_style]),
                                    top=Side(style=border_styles[xf.border.top_line_style]),
                                    bottom=Side(style=border_styles[xf.border.bottom_line_style]))
                    cell_xlsx.border = border

    # 保存为.xlsx文件
    _xlsx_file = p_file_path + 'x'
    workbook_xlsx.save(_xlsx_file)
    workbook_xls.release_resources()
    return _xlsx_file


def table_format(p_dict, p_data_frame):
    _json = {"row": 5, "column": 6}
    print(type(_json))
